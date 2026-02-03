#!/usr/bin/env python3
"""
掃描檔 PDF 定點數據擷取與 Excel 自動化 (OCR-to-Excel Automation)
根據 config.json 設定的像素座標裁切區域並進行 OCR，輸出至 Excel。
"""

import json
import os
import re
from pathlib import Path

import pandas as pd
import pytesseract
from openpyxl.utils import column_index_from_string
from pdf2image import convert_from_path
from PIL import Image


def load_config(path: str = "config.json") -> dict:
    """讀取 JSON 設定檔，取得輸入 PDF、輸出路徑與欄位座標。"""
    config_path = Path(path)
    if not config_path.exists():
        raise FileNotFoundError(f"設定檔不存在: {path}")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def preprocess_image(image: Image.Image) -> Image.Image:
    """
    影像前處理：轉灰階 -> 二值化 -> 去雜訊。
    讓數字變黑、背景變白，提升 OCR 準確率。
    """
    # 轉灰階
    gray = image.convert("L")
    # 二值化：Otsu 或固定閾值，此處用 0/255 二值化
    threshold = 150
    binary = gray.point(lambda p: 255 if p > threshold else 0, mode="1")
    # 轉回 L 模式供 pytesseract 使用
    result = binary.convert("L")
    return result


def extract_data_from_pdf(config: dict) -> list[dict]:
    """
    依設定將 PDF 轉為圖片，對每頁依欄位座標裁切、前處理、OCR，回傳 List[dict]。
    """
    input_pdf = config.get("input_pdf")
    if not input_pdf or not Path(input_pdf).exists():
        raise FileNotFoundError(f"找不到 PDF 檔案: {input_pdf}")

    # 1. PDF 轉為高解析度圖片列表 (DPI >= 300)
    images = convert_from_path(input_pdf, dpi=300)
    extracted_data = []

    # 2. 遍歷每一頁
    for page_num, img in enumerate(images, start=1):
        page_record: dict = {"Page": page_num}

        # 3. 遍歷使用者設定的每個欄位
        for field in config.get("fields", []):
            # 若欄位指定特定頁，可依 page_index 過濾（此處支援 "all"）
            page_index = field.get("page_index", "all")
            if page_index != "all" and page_num != page_index:
                continue

            coords = field["box"]
            # box: [left, top, right, bottom] 像素
            left, top, right, bottom = coords[0], coords[1], coords[2], coords[3]
            cropped_img = img.crop((left, top, right, bottom))

            clean_img = preprocess_image(cropped_img)

            # OCR：單行；可依欄位設定 whitelist（空字串 = 不限制，辨識全部字元，利於血色素等）
            whitelist = field.get("ocr_whitelist", "0123456789.")
            lang = field.get("lang", "eng")
            tesseract_config = "--psm 7"
            if whitelist:
                tesseract_config += f" -c tessedit_char_whitelist={whitelist}"
            text = pytesseract.image_to_string(
                clean_img,
                config=tesseract_config,
                lang=lang,
            )
            value = text.strip()
            page_record[field["name"]] = value

        extracted_data.append(page_record)

    return extracted_data


def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """將儲存格參照如 'B2' 解析為 (row, col_idx)，col_idx 為 1-based。"""
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref.strip())
    if not match:
        raise ValueError(f"無效的儲存格參照: {cell_ref}")
    col_letter, row = match.group(1).upper(), int(match.group(2))
    return row, column_index_from_string(col_letter)


def save_to_excel(
    data: list[dict], output_path: str, config: dict | None = None
) -> None:
    """
    將 List[dict] 寫入 .xlsx。
    若 config 中欄位有指定 "cell"（如 "B2"），則寫入該欄、從該列開始，每頁一列。
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    fields = (config or {}).get("fields", [])
    use_cells = any(f.get("cell") for f in fields)

    if use_cells and fields:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("無法建立工作表")
        omit_page = config.get("omit_page_column", False)
        if not omit_page:
            page_cell = config.get("page_cell", "A2")
            page_row, page_col = _parse_cell_ref(page_cell)
            ws.cell(row=page_row - 1, column=page_col, value="Page")
        # 標題列：各欄位名稱寫在對應 cell 的上一列
        for f in fields:
            if not f.get("cell"):
                continue
            r, c = _parse_cell_ref(f["cell"])
            ws.cell(row=r - 1, column=c, value=f["name"])
        # 資料列：直向列出，每欄一個欄位、每頁一列
        for i, record in enumerate(data):
            row_offset = i
            if not omit_page:
                ws.cell(row=page_row + row_offset, column=page_col, value=record.get("Page"))
            for f in fields:
                cell_ref = f.get("cell")
                if not cell_ref:
                    continue
                r, c = _parse_cell_ref(cell_ref)
                ws.cell(row=r + row_offset, column=c, value=record.get(f["name"], ""))
        wb.save(output_path)
    else:
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine="openpyxl")
    print(f"數據已儲存至 {output_path}")


def main() -> None:
    config_path = os.environ.get("OCR_CONFIG", "config.json")
    conf = load_config(config_path)
    result_array = extract_data_from_pdf(conf)
    save_to_excel(result_array, conf["output_excel"], config=conf)


if __name__ == "__main__":
    main()
